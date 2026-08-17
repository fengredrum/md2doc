#!/usr/bin/env python3
"""将 Markdown 文档中的表格转换为 Excel (.xlsx)。

支持两类表格（按文档顺序依次转换，每个表格一个 worksheet）：
    1. HTML <table> —— 保留 colspan/rowspan、<b> 粗体、<th> 表头、
       <br> 换行，以及 style 中的 text-align 对齐方式；
    2. Markdown 管道表格 —— `| a | b |` 形式，首行为表头，
       分隔行 `:---` / `---:` / `:---:` 决定列对齐。

样式对齐旧版 课程大纲转Excel.py：微软雅黑、细边框、表头加粗 + 浅灰底纹、
垂直居中、自动换行，并按内容估算列宽与行高。

用法：
    uv run python scripts/md2xlsx.py [输入.md] [-o 输出.xlsx]

默认输出为与输入同名的 .xlsx。
"""
import argparse
import math
import re
import sys
from html.parser import HTMLParser
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "微软雅黑"
BODY_SIZE = 11
HEADER_FILL = "F2F2F2"  # 表头浅灰底纹；设为 None 可关闭
MIN_COL_WIDTH = 8.0
MAX_COL_WIDTH = 60.0

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------- HTML 表格解析
class _TableParser(HTMLParser):
    """解析 HTML <table>，提取带 colspan/rowspan/粗体/表头/对齐信息的单元格。"""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.rows = []
        self._row = None
        self._cell = None
        self._in_cell = False
        self._in_bold = False

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th") and self._row is not None:
            self._cell = {
                "text": "",
                "colspan": int(a.get("colspan", 1)),
                "rowspan": int(a.get("rowspan", 1)),
                "is_header": (tag == "th"),
                "bold": False,
                "align": _parse_text_align(a.get("style", "")),
            }
            self._in_cell = True
        elif tag == "b" and self._in_cell:
            self._in_bold = True
        elif tag == "br" and self._in_cell:
            self._cell["text"] += "\n"

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._in_cell:
            self._row.append(self._cell)
            self._cell = None
            self._in_cell = False
            self._in_bold = False
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None
        elif tag == "b" and self._in_cell:
            self._in_bold = False

    def handle_data(self, data):
        if self._in_cell and self._cell is not None:
            self._cell["text"] += data
            if self._in_bold:
                self._cell["bold"] = True


def _parse_text_align(style):
    """从 style 属性中提取 text-align 值（left/center/right）。"""
    m = re.search(r"text-align\s*:\s*(left|center|right)", style or "", re.I)
    return m.group(1) if m else None


def _clean(text):
    text = text.replace("\xa0", " ").replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln.strip() for ln in text.split("\n")]
    return "\n".join(ln for ln in lines if ln)


def _html_to_rows(html_table):
    """解析一段 <table>…</table>，返回 rows（list of list of cell dict）。"""
    parser = _TableParser()
    parser.feed(html_table)
    return parser.rows


# ---------------------------------------------------------------- Markdown 管道表格
def _is_pipe_row(line):
    s = line.strip()
    if not s or s.startswith(("```", "~~~", "#", ">")):
        return False
    return s.startswith("|") or s.endswith("|")


def _is_separator_row(line):
    s = line.strip()
    if not s:
        return False
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    if not s:
        return False
    return all(re.fullmatch(r":?-+:?", c.strip()) for c in s.split("|"))


def _split_pipe_row(line):
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [c.strip() for c in s.split("|")]


def _separator_aligns(sep_line):
    s = sep_line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    out = []
    for cell in s.split("|"):
        cell = cell.strip()
        if cell.startswith(":") and cell.endswith(":"):
            out.append("center")
        elif cell.startswith(":"):
            out.append("left")
        elif cell.endswith(":"):
            out.append("right")
        else:
            out.append(None)
    return out


def _parse_pipe_block(block):
    """把一段连续的管道表格行解析为 grid（首行为表头）。"""
    aligns = []
    if len(block) > 1 and _is_separator_row(block[1]):
        aligns = _separator_aligns(block[1])

    rows = []
    for idx, line in enumerate(block):
        if idx == 1 and _is_separator_row(line):
            continue
        cells = _split_pipe_row(line)
        rows.append([
            {
                "text": text,
                "colspan": 1,
                "rowspan": 1,
                "is_header": (idx == 0),
                "bold": False,
                "align": aligns[c] if c < len(aligns) else None,
            }
            for c, text in enumerate(cells)
        ])
    return _rows_to_grid(rows)


# ---------------------------------------------------------------- 网格展开
def _rows_to_grid(rows):
    """把带 colspan/rowspan 的单元格列表展开为二维网格。

    Returns:
        (grid, spans, header, align)
        grid:   二维文本（合并格仅左上角保存文字，其余为 ""）
        spans:  合并区域 [(row, col, rowspan, colspan), ...]
        header: 二维 bool（是否表头单元格）
        align:  二维 str|None（显式对齐，来自 text-align 或分隔行）
    """
    num_rows = len(rows)
    num_cols = max(
        (sum(max(1, c.get("colspan", 1)) for c in row) for row in rows),
        default=0,
    )
    grid = [[None] * num_cols for _ in range(num_rows)]
    header = [[False] * num_cols for _ in range(num_rows)]
    align = [[None] * num_cols for _ in range(num_rows)]
    spans = []

    for r, cells in enumerate(rows):
        c = 0
        for cell in cells:
            while c < num_cols and grid[r][c] is not None:
                c += 1
            if c >= num_cols:
                break
            text = _clean(cell.get("text", ""))
            is_hdr = bool(cell.get("is_header") or cell.get("bold"))
            cs = max(1, int(cell.get("colspan", 1)))
            rs = max(1, int(cell.get("rowspan", 1)))
            for dc in range(cs):
                for dr in range(rs):
                    rr, cc = r + dr, c + dc
                    if rr < num_rows and cc < num_cols:
                        grid[rr][cc] = text if (dc == 0 and dr == 0) else ""
                        header[rr][cc] = is_hdr
                        align[rr][cc] = cell.get("align")
            if cs > 1 or rs > 1:
                spans.append((r, c, rs, cs))
            c += cs

    for r in range(num_rows):
        for c in range(num_cols):
            if grid[r][c] is None:
                grid[r][c] = ""
    return grid, spans, header, align


def _drop_empty_edge_columns(grid, spans, header, align):
    """去掉首/尾完全空白的列（复刻旧脚本「去掉左侧黑色空列」）。"""
    num_rows = len(grid)
    num_cols = len(grid[0]) if grid else 0

    def col_empty(c):
        return all(grid[r][c] in (None, "") for r in range(num_rows))

    first = 0
    while first < num_cols and col_empty(first):
        first += 1
    last = num_cols - 1
    while last >= first and col_empty(last):
        last -= 1

    if first == 0 and last == num_cols - 1:
        return grid, spans, header, align

    new_grid = [row[first:last + 1] for row in grid]
    new_header = [row[first:last + 1] for row in header]
    new_align = [row[first:last + 1] for row in align]
    new_cols = last - first + 1

    new_spans = []
    for (r, c, rs, cs) in spans:
        nc = c - first
        if nc < 0 or nc >= new_cols:
            continue
        ncs = min(cs, new_cols - nc)
        if ncs <= 0:
            continue
        new_spans.append((r, nc, rs, ncs))
    return new_grid, new_spans, new_header, new_align


# ---------------------------------------------------------------- 表格提取
def _find_preceding_heading(md_text, pos):
    """返回 pos 之前最近的一个 Markdown 标题文本（清洗后），无则 None。"""
    before = md_text[:pos]
    matches = re.findall(r"(?m)^\s{0,3}#{1,6}\s+(.+?)\s*$", before)
    if not matches:
        return None
    title = re.sub(r"[`*_#\[\]()]", "", matches[-1]).strip()
    return title or None


def extract_tables(md_text):
    """提取 md 中所有表格（HTML + 管道），按文档顺序返回。

    Returns:
        list of dict: {"grid", "spans", "header", "align", "title", "start"}
    """
    # 统一换行符，保证后续行偏移计算一致
    md_text = md_text.replace("\r\n", "\n").replace("\r", "\n")

    entries = []
    html_re = re.compile(r"<table[\s>][\s\S]*?</table>", re.IGNORECASE)
    html_matches = list(html_re.finditer(md_text))

    for m in html_matches:
        rows = _html_to_rows(m.group(0))
        grid, spans, header, align = _rows_to_grid(rows)
        grid, spans, header, align = _drop_empty_edge_columns(
            grid, spans, header, align
        )
        entries.append({
            "grid": grid,
            "spans": spans,
            "header": header,
            "align": align,
            "title": _find_preceding_heading(md_text, m.start()),
            "start": m.start(),
        })

    entries.extend(_extract_pipe_tables(md_text, [(m.start(), m.end()) for m in html_matches]))

    entries.sort(key=lambda e: e["start"])
    return entries


def _extract_pipe_tables(md_text, html_ranges):
    lines = md_text.splitlines()
    offsets = []
    pos = 0
    for ln in lines:
        offsets.append(pos)
        pos += len(ln) + 1

    def in_html(idx):
        off = offsets[idx]
        return any(s <= off < e for s, e in html_ranges)

    entries = []
    i = 0
    n = len(lines)
    while i < n:
        if in_html(i):
            i += 1
            continue
        if not _is_pipe_row(lines[i]) or i + 1 >= n or not _is_separator_row(lines[i + 1]):
            i += 1
            continue

        j = i
        block = []
        while j < n and _is_pipe_row(lines[j]):
            block.append(lines[j])
            j += 1

        grid, spans, header, align = _parse_pipe_block(block)
        entries.append({
            "grid": grid,
            "spans": spans,
            "header": header,
            "align": align,
            "title": _find_preceding_heading(md_text, offsets[i]),
            "start": offsets[i],
        })
        i = j
    return entries


# ---------------------------------------------------------------- 尺寸估算
def _vwidth(s):
    """字符串显示宽度：中日韩/全角标点/箭头按 2，其余按 1。"""
    w = 0
    for ch in s:
        o = ord(ch)
        if o >= 0x2E80 or 0x2190 <= o <= 0x21FF:
            w += 2
        else:
            w += 1
    return w


def _lines(text, capacity):
    return sum(max(1, math.ceil(_vwidth(ln) / capacity)) for ln in text.split("\n"))


def _height(text, width, font_size):
    per_line = font_size * 2.0
    return max(font_size * 1.7, _lines(text, width) * per_line + 8)


def _estimate_col_widths(grid, spans):
    """按非合并单元格内容估算列宽；合并单元格不单独撑宽列。

    横向合并格（colspan>1）的文本会跨列自动换行，不参与单列宽度计算，
    以免标题/简介等整行合并格把各列撑到上限。
    """
    num_cols = len(grid[0]) if grid else 0
    merged_topleft = {(r, c) for (r, c, rs, cs) in spans if cs > 1}
    widths = [MIN_COL_WIDTH] * num_cols

    for c in range(num_cols):
        mw = 0
        for r in range(len(grid)):
            if (r, c) in merged_topleft:
                continue
            for ln in grid[r][c].split("\n"):
                mw = max(mw, _vwidth(ln))
        widths[c] = min(max(mw + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
    return widths


def _estimate_row_heights(grid, spans, col_widths, font_size):
    num_rows = len(grid)
    heights = [font_size * 1.7] * num_rows
    span_map = {(r, c): (rs, cs) for (r, c, rs, cs) in spans}

    # 第一遍：普通格与横向合并格决定各自行高
    for r in range(num_rows):
        for c, text in enumerate(grid[r]):
            if not text:
                continue
            rs, cs = span_map.get((r, c), (1, 1))
            if cs > 1:
                eff = sum(col_widths[c:c + cs])
                heights[r] = max(heights[r], _height(text, eff, font_size))
            elif rs > 1:
                continue  # 纵向合并格交给第二遍
            else:
                heights[r] = max(heights[r], _height(text, col_widths[c], font_size))

    # 第二遍：纵向合并格，确保跨行总高足够
    for (r, c, rs, cs) in spans:
        if cs > 1:
            continue
        text = grid[r][c]
        need = _height(text, col_widths[c], font_size)
        total = sum(heights[r:r + rs])
        if need > total:
            heights[r + rs - 1] += need - total
    return heights


# ---------------------------------------------------------------- 生成 Excel
def _style_cell(cell, is_hdr, align, font_name, size):
    cell.font = Font(name=font_name, size=size, bold=is_hdr)
    cell.alignment = Alignment(
        horizontal=align or ("center" if is_hdr else "left"),
        vertical="center",
        wrap_text=True,
    )
    cell.border = BORDER
    if is_hdr and HEADER_FILL:
        cell.fill = PatternFill(fill_type="solid", start_color=HEADER_FILL,
                                end_color=HEADER_FILL)


def _sanitize_sheet_name(name, idx):
    if name:
        name = re.sub(r"[\[\]:*?/\\]", " ", name)
        name = re.sub(r"\s+", " ", name).strip()
    if not name:
        name = f"表格{idx + 1}"
    return name[:31]


def _fill_sheet(ws, table, font_name, body_size):
    grid, spans, header, align = (
        table["grid"], table["spans"], table["header"], table["align"],
    )
    num_rows = len(grid)
    num_cols = len(grid[0]) if grid else 0
    if num_cols == 0:
        return

    col_widths = _estimate_col_widths(grid, spans)
    row_heights = _estimate_row_heights(grid, spans, col_widths, body_size)

    for c, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(c + 1)].width = w
    for r, h in enumerate(row_heights):
        ws.row_dimensions[r + 1].height = h

    # 写入值
    for r in range(num_rows):
        for c in range(num_cols):
            if grid[r][c]:
                ws.cell(row=r + 1, column=c + 1, value=grid[r][c])

    # 合并单元格
    for (r, c, rs, cs) in spans:
        if rs > 1 or cs > 1:
            ws.merge_cells(
                start_row=r + 1, start_column=c + 1,
                end_row=r + rs, end_column=c + cs,
            )

    # 统一样式（合并后对所有单元格，含 MergedCell，应用字体/边框/对齐/底纹）
    for r in range(num_rows):
        for c in range(num_cols):
            cell = ws.cell(row=r + 1, column=c + 1)
            _style_cell(cell, header[r][c], align[r][c], font_name, body_size)


def build_workbook(tables, font_name=FONT, body_size=BODY_SIZE):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used = set()
    for i, table in enumerate(tables):
        base = _sanitize_sheet_name(table.get("title"), i)
        name, n = base, 0
        while name in used:
            n += 1
            name = f"{base[:28]}-{n}"
        used.add(name)
        ws = wb.create_sheet(title=name)
        _fill_sheet(ws, table, font_name, body_size)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="表格1")
        _fill_sheet(ws, {"grid": [[""]], "spans": [], "header": [[False]], "align": [[None]]},
                    font_name, body_size)
    return wb


# ---------------------------------------------------------------- CLI
def main():
    parser = argparse.ArgumentParser(
        description="将 Markdown 文档中的表格转换为 Excel (.xlsx)",
    )
    parser.add_argument("input", nargs="?", default=None, help="输入的 Markdown 文件路径")
    parser.add_argument("-o", "--output", default=None, help="输出 .xlsx 文件路径")
    args = parser.parse_args()

    if args.input is None:
        # 无参数时回退到同目录下的课程三示例文件
        root = Path(__file__).resolve().parent.parent
        src = root / "测试用例" / "AI场景落地全流程实战.md"
    else:
        src = Path(args.input)

    if not src.exists():
        print(f"错误：输入文件不存在: {src}", file=sys.stderr)
        sys.exit(1)

    out = Path(args.output) if args.output else src.with_suffix(".xlsx")
    tables = extract_tables(src.read_text(encoding="utf-8"))
    if not tables:
        print("未在文档中找到任何表格（HTML 或 Markdown 管道表格）。", file=sys.stderr)
        sys.exit(1)

    wb = build_workbook(tables)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已生成：{out}（{len(tables)} 个表格）")


if __name__ == "__main__":
    main()
